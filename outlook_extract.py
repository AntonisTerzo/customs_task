import os
import gc
import ctypes
import threading
import traceback
import tkinter as tk
from tkinter import scrolledtext, messagebox
import pythoncom
import win32com.client
from openpyxl import Workbook

FORMULA_CHARS = {"=", "+", "-", "@", "\t", "\r"}

PGTS_SHARED_MAILBOX = "nlcustomsskg.kngs / Kuehne+Nagel / SKG"          
PGTS_SUBFOLDER      = "PGTS"    

BE_SHARED_MAILBOX   = "belux.customs"     

BE_PATHS = [
    (["1. Control Tower", "BEBRU", "BEBRU Export"], "BEBRU Export"),
    (["1.2 Control Tower Thessaloniki", "BEBRU 3rdparty ECS"],         "ECS"),
    (["POA Registration"],                       "POA"),
]

# ── Shared helpers ────────────────────────────────────────────────────────────

def get_display_name():
    """Return the user's Windows display name, trimmed and safe."""
    try:
        GetUserNameEx = ctypes.windll.secur32.GetUserNameExW
        NameDisplay = 3
        size = ctypes.pointer(ctypes.c_ulong(0))
        GetUserNameEx(NameDisplay, None, size)
        name_buffer = ctypes.create_unicode_buffer(size.contents.value)
        GetUserNameEx(NameDisplay, name_buffer, size)
        username = name_buffer.value or ""
        if "/" in username:
            username = username.split("/")[0].strip()
        if not username:
            username = os.environ.get("USERNAME", "User")
        return username
    except Exception:
        return os.environ.get("USERNAME", "User")

def sanitize_subject(subject):
    """Replace a leading formula-injection character with underscore."""
    if subject and subject[0] in FORMULA_CHARS:
        subject = "_" + subject[1:]
    return subject

def resolve_output_path(out_dir, base_name):
    """
    Return a unique file path inside out_dir.
    If <base_name>.xlsx already exists, tries <base_name>(1).xlsx, etc.
    """
    candidate = os.path.join(out_dir, f"{base_name}.xlsx")
    if not os.path.exists(candidate):
        return candidate
    counter = 1
    while True:
        candidate = os.path.join(out_dir, f"{base_name}({counter}).xlsx")
        if not os.path.exists(candidate):
            return candidate
        counter += 1

def read_folder_emails(folder, log):
    """Return (rows, skipped) for a given Outlook folder."""
    rows = []
    skipped = 0
    items = folder.Items
    items.Sort("[ReceivedTime]", True)
    for i in range(items.Count):
        try:
            item = items.Item(i + 1)
            if item.Class == 43:              # olMailItem = 43
                subject      = sanitize_subject(item.Subject or "(no subject)")
                received     = item.ReceivedTime
                datetime_str = received.strftime("%Y/%m/%d %H:%M:%S")
                rows.append((subject, datetime_str))
        except Exception as exc:
            skipped += 1
            log(f"  WARNING: Skipped item #{i + 1} ({type(exc).__name__}: {exc})")
            continue
    return rows, skipped

def connect_outlook():
    """Return (outlook, mapi, error_message)."""
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mapi = outlook.GetNamespace("MAPI")
        mapi.Logon()
        return outlook, mapi, None
    except Exception as exc:
        return None, None, f"Could not connect to Outlook:\n{exc}"

def release_com():
    """Force garbage collection to release lingering COM references."""
    gc.collect()

# ── Shared mailbox navigation ─────────────────────────────────────────────────

def find_shared_store(mapi, store_name):
    """
    Find a top-level shared mailbox by its display name.
    Returns (store_root_folder, error_message).
    """
    target = store_name.strip().lower()
    try:
        for i in range(mapi.Folders.Count):
            store = mapi.Folders.Item(i + 1)
            if store.Name.strip().lower() == target:
                return store, None
    except Exception as exc:
        return None, f"Error listing mailboxes:\n{exc}"
    return None, (f"Shared mailbox '{store_name}' was not found.\n"
                  f"Make sure it is added to your Outlook profile.")

def find_subfolder(parent_folder, target_name):
    """Return the direct subfolder matching name, or None."""
    target = target_name.strip().lower()
    for i in range(parent_folder.Folders.Count):
        f = parent_folder.Folders.Item(i + 1)
        if f.Name.strip().lower() == target:
            return f
    return None

def walk_path(root_folder, path_parts):
    """
    Walk a list of folder names starting from root_folder.
    Returns (folder, error_message).
    """
    current = root_folder
    walked = []
    for name in path_parts:
        nxt = find_subfolder(current, name)
        if nxt is None:
            walked_str = " / ".join(walked) if walked else "(root)"
            return None, f"Subfolder '{name}' not found under '{walked_str}'."
        current = nxt
        walked.append(name)
    return current, None

# ── PGTS task ─────────────────────────────────────────────────────────────────

def get_pgts_folder(mapi):
    
    store, err = find_shared_store(mapi, PGTS_SHARED_MAILBOX)
    if store is None:
        return None, err

    inbox = find_subfolder(store, "Inbox")
    if inbox is None:
        return None, f"'Inbox' was not found inside shared mailbox '{PGTS_SHARED_MAILBOX}'."

    pgts = find_subfolder(inbox, PGTS_SUBFOLDER)
    if pgts is None:
        return None, (f"Subfolder '{PGTS_SUBFOLDER}' was not found inside "
                      f"'{PGTS_SHARED_MAILBOX}' / Inbox.")

    return pgts, None

def run_pgts_task(log):
    log("Starting PGTS task...")
    log("Connecting to Outlook...")

    outlook = mapi = pgts_folder = None
    try:
        outlook, mapi, err = connect_outlook()
        if err:
            log(f"ERROR: {err}")
            messagebox.showerror("Error", err)
            return False, None

        log(f"Locating '{PGTS_SHARED_MAILBOX}' / Inbox / {PGTS_SUBFOLDER}...")
        pgts_folder, err = get_pgts_folder(mapi)
        if pgts_folder is None:
            log(f"ERROR: {err}")
            messagebox.showerror("Folder Not Found", err)
            return False, None

        log("Reading emails...")
        try:
            rows, skipped = read_folder_emails(pgts_folder, log)
        except Exception as exc:
            log(f"ERROR: Failed to read emails: {exc}")
            log(traceback.format_exc())
            messagebox.showerror("Error", f"Failed to read emails:\n{exc}")
            return False, None

        if not rows:
            log("Folder is empty.")
            messagebox.showwarning(
                "Empty Folder",
                f"The folder '{PGTS_SHARED_MAILBOX} / Inbox / {PGTS_SUBFOLDER}' "
                "was found but contains no emails."
            )
            return False, None

        log(f"Found {len(rows)} email(s). Writing Excel file...")
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out_dir   = os.path.join(downloads, "PGTS")
        os.makedirs(out_dir, exist_ok=True)
        out_path  = resolve_output_path(out_dir, "outlook_emails")

        wb = Workbook()
        ws = wb.active
        ws.title = "Emails"
        ws.append(["Subject", "Received"])
        for row in rows:
            ws.append(list(row))
        wb.save(out_path)

        log(f"Saved to: {out_path}")
        if skipped:
            log(f"{skipped} item(s) skipped (see warnings above).")
        log("PGTS task complete.")

        summary = f"{len(rows)} email(s) exported to:\n{out_path}"
        if skipped:
            summary += f"\n\n{skipped} item(s) were skipped (see log)."
        messagebox.showinfo("PGTS Complete", summary)
        return True, out_dir
    finally:
        pgts_folder = None
        mapi = None
        outlook = None
        release_com()

# ── BE task ───────────────────────────────────────────────────────────────────

def run_be_task(log):
    log("Starting BE task...")
    log("Connecting to Outlook...")

    outlook = mapi = store = None
    try:
        outlook, mapi, err = connect_outlook()
        if err:
            log(f"ERROR: {err}")
            messagebox.showerror("Error", err)
            return False, None

        log(f"Locating shared mailbox '{BE_SHARED_MAILBOX}'...")
        store, err = find_shared_store(mapi, BE_SHARED_MAILBOX)
        if store is None:
            log(f"ERROR: {err}")
            messagebox.showerror("Shared Mailbox Not Found", err)
            return False, None

        wb = Workbook()
        wb.remove(wb.active)          # remove default empty sheet

        report_lines = []
        total_rows = 0
        total_skipped = 0

        for path_parts, sheet_name in BE_PATHS:
            path_display = f"{BE_SHARED_MAILBOX} / " + " / ".join(path_parts)
            log(f"Processing '{path_display}'...")
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Subject", "Received"])

            folder, walk_err = walk_path(store, path_parts)
            if folder is None:
                log(f"  NOT FOUND: {walk_err}")
                report_lines.append(f"- {path_display}: NOT FOUND (empty sheet created)")
                continue

            try:
                rows, skipped = read_folder_emails(folder, log)
            except Exception as exc:
                log(f"  ERROR reading: {exc}")
                log(traceback.format_exc())
                report_lines.append(f"- {path_display}: ERROR reading ({exc})")
                folder = None
                continue

            for row in rows:
                ws.append(list(row))

            total_rows += len(rows)
            total_skipped += skipped

            if not rows:
                log("  Empty - only headers written")
                report_lines.append(f"- {path_display}: EMPTY (headers only)")
            else:
                note = f"- {path_display}: {len(rows)} email(s)"
                if skipped:
                    note += f", {skipped} skipped"
                report_lines.append(note)
                log(f"  {len(rows)} email(s) written" +
                    (f", {skipped} skipped" if skipped else ""))

            folder = None       # release per-folder COM reference

        log("Writing Excel file...")
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out_dir   = os.path.join(downloads, "BE")
        os.makedirs(out_dir, exist_ok=True)
        out_path  = resolve_output_path(out_dir, "be_emails")
        wb.save(out_path)

        log(f"Saved to: {out_path}")
        log("BE task complete.")

        summary  = "BE export complete.\n\n"
        summary += "\n".join(report_lines)
        summary += f"\n\nTotal: {total_rows} email(s) exported"
        if total_skipped:
            summary += f", {total_skipped} skipped"
        summary += f"\nSaved to:\n{out_path}"
        messagebox.showinfo("BE Complete", summary)
        return True, out_dir
    finally:
        store = None
        mapi = None
        outlook = None
        release_com()

# ── GUI ───────────────────────────────────────────────────────────────────────

class OutlookExporterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Outlook -> Excel")
        self.root.geometry("700x520")

        username = get_display_name()

        welcome_label = tk.Label(
            root,
            text=f"Welcome back {username}",
            font=("Arial", 18, "bold"),
            fg="#2C3E50"
        )
        welcome_label.pack(pady=20)

        buttons_frame = tk.Frame(root)
        buttons_frame.pack(pady=10)

        self.pgts_button = tk.Button(
            buttons_frame,
            text="PGTS",
            command=self.start_pgts,
            font=("Arial", 12, "bold"),
            bg="#5DADE2",
            fg="white",
            width=15,
            height=1,
            relief="raised",
            borderwidth=3,
            cursor="hand2"
        )
        self.pgts_button.grid(row=0, column=0, padx=10)

        self.be_button = tk.Button(
            buttons_frame,
            text="BE",
            command=self.start_be,
            font=("Arial", 12, "bold"),
            bg="#5DADE2",
            fg="white",
            width=15,
            height=1,
            relief="raised",
            borderwidth=3,
            cursor="hand2"
        )
        self.be_button.grid(row=0, column=1, padx=10)

        log_label = tk.Label(root, text="Log Output:", font=("Arial", 10, "bold"))
        log_label.pack(pady=(20, 5))

        self.log_text = scrolledtext.ScrolledText(
            root, width=75, height=18, state="disabled")
        self.log_text.pack(pady=10)

    def log(self, message):
        self.root.after(0, self._append_log, message)

    def _append_log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")

    def clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state="disabled")

    def disable_buttons(self):
        self.pgts_button.config(state="disabled")
        self.be_button.config(state="disabled")

    def enable_buttons(self):
        self.pgts_button.config(state="normal", text="PGTS")
        self.be_button.config(state="normal", text="BE")

    def start_pgts(self):
        self.disable_buttons()
        self.pgts_button.config(text="Processing...")
        self.clear_log()
        thread = threading.Thread(target=self._pgts_worker, daemon=True)
        thread.start()

    def _pgts_worker(self):
        pythoncom.CoInitialize()
        try:
            success, folder_path = run_pgts_task(self.log)
            if success and folder_path:
                self.root.after(0, os.startfile, folder_path)
        except Exception as exc:
            self.log(f"ERROR: {exc}")
            self.log(traceback.format_exc())
            self.root.after(0, messagebox.showerror,
                            "Error", f"An error occurred: {exc}")
        finally:
            pythoncom.CoUninitialize()
            self.root.after(0, self.enable_buttons)

    def start_be(self):
        self.disable_buttons()
        self.be_button.config(text="Processing...")
        self.clear_log()
        thread = threading.Thread(target=self._be_worker, daemon=True)
        thread.start()

    def _be_worker(self):
        pythoncom.CoInitialize()
        try:
            success, folder_path = run_be_task(self.log)
            if success and folder_path:
                self.root.after(0, os.startfile, folder_path)
        except Exception as exc:
            self.log(f"ERROR: {exc}")
            self.log(traceback.format_exc())
            self.root.after(0, messagebox.showerror,
                            "Error", f"An error occurred: {exc}")
        finally:
            pythoncom.CoUninitialize()
            self.root.after(0, self.enable_buttons)


if __name__ == "__main__":
    root = tk.Tk()
    app = OutlookExporterGUI(root)
    root.mainloop()
